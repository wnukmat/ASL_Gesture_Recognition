import sys, os
import openpyxl
import urllib2

def get_vid(url, word, num):
	file_name = url.split('/')[-1]
	u = urllib2.urlopen(url)
	with open('data/' + word + '/' + str(num) + '.mov','wb') as output:
		output.write(u.read())
	
def main():
	try:											#check for data directory and create if needed
		os.stat('data/')
	except:
		os.mkdir('data/')
				
	wb = openpyxl.load_workbook('test.xlsx')		
	ws = wb.get_sheet_by_name('Sheet1')

	for i in range(3,13236):
		word = ws.cell(row=i, column=2).value		#check for new word
		
		if(word == ' '):							#disregard variations of words
			pass
		elif(word != None):							#if new word
			word = word.replace("/", "-")
			dir = word
			example = 1								#set to first example of word
			try:
				os.stat('data/' + dir)				#check for directory associated with word
			except:
				os.mkdir('data/' + dir) 			#create if needed
		else:
			url = str(ws.cell(row=i, column=12).value[12:-9])	#get hyperlink
			get_vid(url, dir, example)				#download video
			example += 1							#increment to next example of word
			

if __name__ == "__main__":
	main()